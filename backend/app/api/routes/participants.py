from io import BytesIO
from uuid import UUID

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
import openpyxl
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.api.deps import get_current_user
from app.db.session import get_db
from app.models.evaluation import Evaluation
from app.models.participant import EvaluatorAssignment, Participant
from app.schemas.auth import CurrentUser
from app.schemas.participant import (
    AssignmentCreate,
    AssignmentRead,
    CSVUploadResult,
    ParticipantCreate,
    ParticipantRead,
    ParticipantUpdate,
)
from app.services.csv_service import parse_participants_csv

router = APIRouter()


def _ensure_evaluation(db: Session, company_id: UUID, evaluation_id: UUID) -> None:
    evaluation = db.get(Evaluation, evaluation_id)
    if not evaluation or evaluation.company_id != company_id:
        raise HTTPException(status_code=404, detail="Evaluation not found")


@router.post("/{evaluation_id}/participants", response_model=ParticipantRead)
def create_participant(
    evaluation_id: UUID,
    payload: ParticipantCreate,
    current_user: CurrentUser = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> Participant:
    _ensure_evaluation(db, current_user.company_id, evaluation_id)
    participant = Participant(
        company_id=current_user.company_id,
        evaluation_id=evaluation_id,
        email=payload.email.lower(),
        full_name=payload.full_name,
        role=payload.role,
    )
    db.add(participant)
    db.flush()  # Generar ID del participante
    
    # Auto-crear asignación de autoevaluación (self)
    auto_assignment = EvaluatorAssignment(
        company_id=current_user.company_id,
        evaluation_id=evaluation_id,
        evaluatee_id=participant.id,
        evaluator_id=participant.id,
        relationship="self",
        weight=1.0,
    )
    db.add(auto_assignment)
    db.commit()
    db.refresh(participant)
    return participant


@router.post("/{evaluation_id}/participants/csv", response_model=CSVUploadResult)
async def upload_participants_csv(
    evaluation_id: UUID,
    file: UploadFile = File(...),
    current_user: CurrentUser = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> CSVUploadResult:
    _ensure_evaluation(db, current_user.company_id, evaluation_id)
    rows = parse_participants_csv((await file.read()).decode("utf-8-sig"))
    created = 0
    skipped = 0
    for row in rows:
        email = row["email"].lower()
        exists = db.scalar(
            select(Participant).where(
                Participant.company_id == current_user.company_id,
                Participant.evaluation_id == evaluation_id,
                Participant.email == email,
            )
        )
        if exists:
            skipped += 1
            continue
        
        participant = Participant(
            company_id=current_user.company_id,
            evaluation_id=evaluation_id,
            email=email,
            full_name=row["full_name"],
            role=row["role"],
        )
        db.add(participant)
        db.flush()  # Generar ID
        
        # Auto-crear autoevaluación
        auto_assignment = EvaluatorAssignment(
            company_id=current_user.company_id,
            evaluation_id=evaluation_id,
            evaluatee_id=participant.id,
            evaluator_id=participant.id,
            relationship="self",
            weight=1.0,
        )
        db.add(auto_assignment)
        created += 1
    db.commit()
    return CSVUploadResult(created=created, skipped=skipped)


@router.get("/{evaluation_id}/participants", response_model=list[ParticipantRead])
def list_participants(
    evaluation_id: UUID,
    current_user: CurrentUser = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> list[Participant]:
    _ensure_evaluation(db, current_user.company_id, evaluation_id)
    return list(
        db.scalars(
            select(Participant).where(
                Participant.company_id == current_user.company_id,
                Participant.evaluation_id == evaluation_id,
            )
        )
    )


@router.patch("/{evaluation_id}/participants/{participant_id}", response_model=ParticipantRead)
def update_participant(
    evaluation_id: UUID,
    participant_id: UUID,
    payload: ParticipantUpdate,
    current_user: CurrentUser = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> Participant:
    _ensure_evaluation(db, current_user.company_id, evaluation_id)
    participant = db.get(Participant, participant_id)
    if (
        not participant
        or participant.company_id != current_user.company_id
        or participant.evaluation_id != evaluation_id
    ):
        raise HTTPException(status_code=404, detail="Participant not found")

    updates = payload.model_dump(exclude_unset=True)
    if "email" in updates and updates["email"] is not None:
        updates["email"] = updates["email"].lower()
    for field, value in updates.items():
        setattr(participant, field, value)
    db.commit()
    db.refresh(participant)
    return participant


@router.delete("/{evaluation_id}/participants/{participant_id}", status_code=204)
def delete_participant(
    evaluation_id: UUID,
    participant_id: UUID,
    current_user: CurrentUser = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> None:
    _ensure_evaluation(db, current_user.company_id, evaluation_id)
    participant = db.get(Participant, participant_id)
    if (
        not participant
        or participant.company_id != current_user.company_id
        or participant.evaluation_id != evaluation_id
    ):
        raise HTTPException(status_code=404, detail="Participant not found")

    db.delete(participant)
    db.commit()


@router.post("/{evaluation_id}/assignments", response_model=AssignmentRead)
def create_assignment(
    evaluation_id: UUID,
    payload: AssignmentCreate,
    current_user: CurrentUser = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> EvaluatorAssignment:
    _ensure_evaluation(db, current_user.company_id, evaluation_id)
    evaluatee = db.get(Participant, payload.evaluatee_id)
    evaluator = db.get(Participant, payload.evaluator_id)
    if (
        not evaluatee
        or not evaluator
        or evaluatee.company_id != current_user.company_id
        or evaluator.company_id != current_user.company_id
    ):
        raise HTTPException(status_code=404, detail="Participant not found")

    # Evitar asignaciones duplicadas
    exists = db.scalar(
        select(EvaluatorAssignment).where(
            EvaluatorAssignment.company_id == current_user.company_id,
            EvaluatorAssignment.evaluation_id == evaluation_id,
            EvaluatorAssignment.evaluatee_id == payload.evaluatee_id,
            EvaluatorAssignment.evaluator_id == payload.evaluator_id,
        )
    )
    if exists:
        raise HTTPException(status_code=400, detail="Esta asignación ya existe")

    assignment = EvaluatorAssignment(
        company_id=current_user.company_id,
        evaluation_id=evaluation_id,
        evaluatee_id=payload.evaluatee_id,
        evaluator_id=payload.evaluator_id,
        relationship=payload.relationship,
        weight=payload.weight,
    )
    db.add(assignment)
    db.commit()
    db.refresh(assignment)
    return assignment


@router.get("/{evaluation_id}/assignments", response_model=list[AssignmentRead])
def list_assignments(
    evaluation_id: UUID,
    current_user: CurrentUser = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> list[EvaluatorAssignment]:
    _ensure_evaluation(db, current_user.company_id, evaluation_id)
    return list(
        db.scalars(
            select(EvaluatorAssignment).where(
                EvaluatorAssignment.company_id == current_user.company_id,
                EvaluatorAssignment.evaluation_id == evaluation_id,
            )
        )
    )


@router.get("/{evaluation_id}/participants/import-template")
def download_participants_template(
    evaluation_id: UUID,
    current_user: CurrentUser = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> StreamingResponse:
    _ensure_evaluation(db, current_user.company_id, evaluation_id)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Participantes y Relaciones"

    headers = [
        "Email Evaluado", 
        "Nombre Evaluado", 
        "Rol Evaluado", 
        "Email Evaluador", 
        "Relacion", 
        "Peso Relacion (%)"
    ]
    ws.append(headers)

    # Filas de ejemplo
    ws.append([
        "laura@abcforchange.com",
        "Laura Esparza",
        "Representante de Ventas",
        "mguinea@mindoneers.mx",
        "Líder directo",
        100
    ])
    ws.append([
        "laura@abcforchange.com",
        "Laura Esparza",
        "Representante de Ventas",
        "pedro@abcforchange.com",
        "Par",
        50
    ])

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    out = BytesIO()
    wb.save(out)
    out.seek(0)

    filename = "modelo_participantes_relaciones.xlsx"
    headers = {
        "Content-Disposition": f'attachment; filename="{filename}"'
    }
    return StreamingResponse(
        out,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers
    )


@router.post("/{evaluation_id}/participants/import")
async def import_participants_xlsx(
    evaluation_id: UUID,
    file: UploadFile = File(...),
    current_user: CurrentUser = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> dict:
    _ensure_evaluation(db, current_user.company_id, evaluation_id)

    try:
        content = await file.read()
        wb = openpyxl.load_workbook(filename=BytesIO(content), data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="El archivo no es un Excel XLSX válido")

    ws = wb.active
    if not ws:
        raise HTTPException(status_code=400, detail="La hoja de cálculo está vacía")

    participants_created = 0
    assignments_created = 0
    assignments_skipped = 0

    local_participant_cache = {}

    # Mapeo de relación en español a código de backend
    rel_mapping = {
        "par": "peer",
        "reporte directo": "direct_report",
        "líder directo": "line_manager",
        "lider directo": "line_manager",
        "líder indirecto": "indirect_manager",
        "lider indirecto": "indirect_manager",
        "cliente externo": "external_client",
        "cliente interno": "internal_client",
    }

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue

        evaluatee_email = str(row[0]).strip().lower()
        evaluatee_name = str(row[1]).strip() if row[1] is not None else ""
        evaluatee_role = str(row[2]).strip() if row[2] is not None else None
        evaluator_email = str(row[3]).strip().lower() if row[3] is not None else None
        relationship_str = str(row[4]).strip().lower() if row[4] is not None else None
        
        try:
            # Si el peso viene como ej. 50 (porcentaje), lo dividimos por 100 -> 0.5
            weight_val = float(row[5]) / 100.0 if row[5] is not None else 1.0
        except (ValueError, TypeError):
            weight_val = 1.0

        if not evaluatee_email or not evaluatee_name:
            continue

        # 1. Obtener o crear participante Evaluado
        evaluatee = local_participant_cache.get(evaluatee_email)
        if not evaluatee:
            evaluatee = db.scalar(
                select(Participant).where(
                    Participant.company_id == current_user.company_id,
                    Participant.evaluation_id == evaluation_id,
                    Participant.email == evaluatee_email,
                )
            )
        
        if not evaluatee:
            evaluatee = Participant(
                company_id=current_user.company_id,
                evaluation_id=evaluation_id,
                email=evaluatee_email,
                full_name=evaluatee_name,
                role=evaluatee_role,
            )
            db.add(evaluatee)
            db.flush()
            participants_created += 1
            local_participant_cache[evaluatee_email] = evaluatee

            # Auto-crear autoevaluación (self) para el evaluado
            auto_self = EvaluatorAssignment(
                company_id=current_user.company_id,
                evaluation_id=evaluation_id,
                evaluatee_id=evaluatee.id,
                evaluator_id=evaluatee.id,
                relationship="self",
                weight=1.0,
            )
            db.add(auto_self)

        # 2. Si hay evaluador, procesar la asignación (relación)
        if evaluator_email and relationship_str:
            relationship = rel_mapping.get(relationship_str, "peer")

            evaluator = local_participant_cache.get(evaluator_email)
            if not evaluator:
                evaluator = db.scalar(
                    select(Participant).where(
                        Participant.company_id == current_user.company_id,
                        Participant.evaluation_id == evaluation_id,
                        Participant.email == evaluator_email,
                    )
                )
            
            if not evaluator:
                evaluator = Participant(
                    company_id=current_user.company_id,
                    evaluation_id=evaluation_id,
                    email=evaluator_email,
                    full_name=evaluator_email.split("@")[0].capitalize(),
                    role=None,
                )
                db.add(evaluator)
                db.flush()
                participants_created += 1
                local_participant_cache[evaluator_email] = evaluator

                # Auto-crear autoevaluación para el evaluador
                auto_self_eval = EvaluatorAssignment(
                    company_id=current_user.company_id,
                    evaluation_id=evaluation_id,
                    evaluatee_id=evaluator.id,
                    evaluator_id=evaluator.id,
                    relationship="self",
                    weight=1.0,
                )
                db.add(auto_self_eval)

            # Verificar si la asignación ya existe
            assignment_exists = db.scalar(
                select(EvaluatorAssignment).where(
                    EvaluatorAssignment.company_id == current_user.company_id,
                    EvaluatorAssignment.evaluation_id == evaluation_id,
                    EvaluatorAssignment.evaluatee_id == evaluatee.id,
                    EvaluatorAssignment.evaluator_id == evaluator.id,
                )
            )

            if assignment_exists:
                assignment_exists.relationship = relationship
                assignment_exists.weight = weight_val
                assignments_skipped += 1
            else:
                new_assignment = EvaluatorAssignment(
                    company_id=current_user.company_id,
                    evaluation_id=evaluation_id,
                    evaluatee_id=evaluatee.id,
                    evaluator_id=evaluator.id,
                    relationship=relationship,
                    weight=weight_val,
                )
                db.add(new_assignment)
                assignments_created += 1

    db.commit()

    return {
        "participants_created": participants_created,
        "assignments_created": assignments_created,
        "assignments_updated_or_skipped": assignments_skipped,
    }
