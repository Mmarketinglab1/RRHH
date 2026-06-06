from io import BytesIO
from uuid import UUID

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
import openpyxl
from sqlalchemy import func, select
from sqlalchemy.orm import Session

from app.api.deps import get_current_user
from app.db.session import get_db
from app.models.evaluation import Competency, Evaluation, Question, QuestionBank
from app.schemas.auth import CurrentUser
from app.schemas.evaluation import (
    CompetencyCreate,
    CompetencyRead,
    CompetencyUpdate,
    QuestionCreate,
    QuestionRead,
    QuestionUpdate,
    QuestionBankRead,
)

router = APIRouter()


def _ensure_evaluation(db: Session, company_id: UUID, evaluation_id: UUID) -> Evaluation:
    evaluation = db.get(Evaluation, evaluation_id)
    if not evaluation or evaluation.company_id != company_id:
        raise HTTPException(status_code=404, detail="Evaluation not found")
    return evaluation


@router.post("/{evaluation_id}/competencies", response_model=CompetencyRead)
def create_competency(
    evaluation_id: UUID,
    payload: CompetencyCreate,
    current_user: CurrentUser = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> Competency:
    _ensure_evaluation(db, current_user.company_id, evaluation_id)
    competency = Competency(
        company_id=current_user.company_id,
        evaluation_id=evaluation_id,
        competency_bank_id=payload.competency_bank_id,
        name=payload.name,
        description=payload.description,
        weight=payload.weight,
    )
    db.add(competency)
    db.commit()
    db.refresh(competency)
    return competency


@router.get("/{evaluation_id}/competencies", response_model=list[CompetencyRead])
def list_competencies(
    evaluation_id: UUID,
    current_user: CurrentUser = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> list[Competency]:
    _ensure_evaluation(db, current_user.company_id, evaluation_id)
    return list(
        db.scalars(
            select(Competency).where(
                Competency.company_id == current_user.company_id,
                Competency.evaluation_id == evaluation_id,
            )
        )
    )


@router.patch("/{evaluation_id}/competencies/{competency_id}", response_model=CompetencyRead)
def update_competency(
    evaluation_id: UUID,
    competency_id: UUID,
    payload: CompetencyUpdate,
    current_user: CurrentUser = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> Competency:
    _ensure_evaluation(db, current_user.company_id, evaluation_id)
    competency = db.get(Competency, competency_id)
    if (
        not competency
        or competency.company_id != current_user.company_id
        or competency.evaluation_id != evaluation_id
    ):
        raise HTTPException(status_code=404, detail="Competency not found")

    updates = payload.model_dump(exclude_unset=True)
    for field, value in updates.items():
        setattr(competency, field, value)
    db.commit()
    db.refresh(competency)
    return competency


@router.delete("/{evaluation_id}/competencies/{competency_id}", status_code=204)
def delete_competency(
    evaluation_id: UUID,
    competency_id: UUID,
    current_user: CurrentUser = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> None:
    _ensure_evaluation(db, current_user.company_id, evaluation_id)
    competency = db.get(Competency, competency_id)
    if (
        not competency
        or competency.company_id != current_user.company_id
        or competency.evaluation_id != evaluation_id
    ):
        raise HTTPException(status_code=404, detail="Competency not found")

    db.delete(competency)
    db.commit()


@router.post("/{evaluation_id}/questions", response_model=QuestionRead)
def create_question(
    evaluation_id: UUID,
    payload: QuestionCreate,
    current_user: CurrentUser = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> Question:
    _ensure_evaluation(db, current_user.company_id, evaluation_id)
    competency = db.get(Competency, payload.competency_id)
    if not competency or competency.company_id != current_user.company_id:
        raise HTTPException(status_code=404, detail="Competency not found")

    tag_s = payload.tag_self or f"P{payload.position or 1} Autoevaluado"
    tag_e = payload.tag_evaluator or f"P{payload.position or 1}B Evaluador"

    question = Question(
        company_id=current_user.company_id,
        evaluation_id=evaluation_id,
        competency_id=payload.competency_id,
        text=payload.text,
        text_self=payload.text_self or payload.text,
        text_evaluator=payload.text_evaluator or payload.text,
        tag_self=tag_s,
        tag_evaluator=tag_e,
        position=payload.position,
        question_type=payload.question_type,
        options=payload.options,
        is_evaluative=payload.is_evaluative,
    )
    db.add(question)

    if payload.save_to_bank:
        stmt = select(QuestionBank).where(
            QuestionBank.company_id == current_user.company_id,
            func.lower(QuestionBank.competency_name) == func.lower(competency.name),
            func.lower(QuestionBank.text) == func.lower(payload.text)
        )
        existing = db.scalars(stmt).first()
        if not existing:
            bank_item = QuestionBank(
                company_id=current_user.company_id,
                competency_name=competency.name,
                text=payload.text,
                text_self=payload.text_self or payload.text,
                text_evaluator=payload.text_evaluator or payload.text,
                tag_self=tag_s,
                tag_evaluator=tag_e,
                question_type=payload.question_type,
                options=payload.options,
                is_evaluative=payload.is_evaluative
            )
            db.add(bank_item)

    db.commit()
    db.refresh(question)
    return question


@router.get("/{evaluation_id}/suggested-questions", response_model=list[QuestionBankRead])
def list_suggested_questions(
    evaluation_id: UUID,
    competency_id: UUID,
    current_user: CurrentUser = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> list[QuestionBank]:
    _ensure_evaluation(db, current_user.company_id, evaluation_id)
    competency = db.get(Competency, competency_id)
    if not competency or competency.company_id != current_user.company_id:
        raise HTTPException(status_code=404, detail="Competency not found")

    # Si la competencia está enlazada al banco de competencias, buscar por ID
    if competency.competency_bank_id:
        stmt = select(QuestionBank).where(
            QuestionBank.competency_bank_id == competency.competency_bank_id,
            (QuestionBank.company_id == None) | (QuestionBank.company_id == current_user.company_id)
        )
        questions = list(db.scalars(stmt).all())
        if questions:
            return questions

    comp_name_lower = competency.name.strip().lower()
    stmt = select(QuestionBank).where(
        func.lower(QuestionBank.competency_name) == comp_name_lower,
        (QuestionBank.company_id == None) | (QuestionBank.company_id == current_user.company_id)
    )
    return list(db.scalars(stmt).all())


@router.get("/{evaluation_id}/questions", response_model=list[QuestionRead])
def list_questions(
    evaluation_id: UUID,
    current_user: CurrentUser = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> list[Question]:
    _ensure_evaluation(db, current_user.company_id, evaluation_id)
    return list(
        db.scalars(
            select(Question)
            .where(
                Question.company_id == current_user.company_id,
                Question.evaluation_id == evaluation_id,
            )
            .order_by(Question.position.asc())
        )
    )


@router.patch("/{evaluation_id}/questions/{question_id}", response_model=QuestionRead)
def update_question(
    evaluation_id: UUID,
    question_id: UUID,
    payload: QuestionUpdate,
    current_user: CurrentUser = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> Question:
    _ensure_evaluation(db, current_user.company_id, evaluation_id)
    question = db.get(Question, question_id)
    if (
        not question
        or question.company_id != current_user.company_id
        or question.evaluation_id != evaluation_id
    ):
        raise HTTPException(status_code=404, detail="Question not found")

    updates = payload.model_dump(exclude_unset=True)
    if "competency_id" in updates:
        competency = db.get(Competency, updates["competency_id"])
        if (
            not competency
            or competency.company_id != current_user.company_id
            or competency.evaluation_id != evaluation_id
        ):
            raise HTTPException(status_code=404, detail="Competency not found")
    for field, value in updates.items():
        setattr(question, field, value)
    db.commit()
    db.refresh(question)
    return question


@router.delete("/{evaluation_id}/questions/{question_id}", status_code=204)
def delete_question(
    evaluation_id: UUID,
    question_id: UUID,
    current_user: CurrentUser = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> None:
    _ensure_evaluation(db, current_user.company_id, evaluation_id)
    question = db.get(Question, question_id)
    if (
        not question
        or question.company_id != current_user.company_id
        or question.evaluation_id != evaluation_id
    ):
        raise HTTPException(status_code=404, detail="Question not found")

    db.delete(question)
    db.commit()


@router.get("/{evaluation_id}/competencies/import-template")
def download_import_template(
    evaluation_id: UUID,
    current_user: CurrentUser = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> StreamingResponse:
    _ensure_evaluation(db, current_user.company_id, evaluation_id)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Competencias y Preguntas"

    # Set headers
    headers = [
        "Competencia", 
        "Descripción de la Competencia", 
        "Peso", 
        "Pregunta", 
        "Tipo de Pregunta", 
        "Opciones (separadas por coma)", 
        "Evaluativa (SI/NO)"
    ]
    ws.append(headers)

    # Example rows
    ws.append([
        "Liderazgo",
        "Habilidad para dirigir, coordinar e impulsar al equipo.",
        1.0,
        "¿El evaluado demuestra iniciativa para liderar nuevos proyectos?",
        "numeric_1_10",
        "",
        "SI"
    ])
    ws.append([
        "Liderazgo",
        "Habilidad para dirigir, coordinar e impulsar al equipo.",
        1.0,
        "¿El evaluado delega responsabilidades de manera equilibrada y clara?",
        "likert",
        "Muy en desacuerdo, En desacuerdo, Neutral, De acuerdo, Muy de acuerdo",
        "SI"
    ])
    ws.append([
        "Trabajo en Equipo",
        "Capacidad de colaborar con otros para lograr metas comunes.",
        1.2,
        "¿El evaluado comparte información y conocimientos de forma abierta?",
        "dicotomic",
        "Sí, No",
        "SI"
    ])
    ws.append([
        "Trabajo en Equipo",
        "Capacidad de colaborar con otros para lograr metas comunes.",
        1.2,
        "¿Qué áreas de mejora técnica sugerirías para el evaluado?",
        "checklist",
        "Puntualidad, Comunicación, Calidad técnica, Organización",
        "NO"
    ])

    # Auto-fit column widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    out = BytesIO()
    wb.save(out)
    out.seek(0)

    filename = "modelo_competencias_preguntas.xlsx"
    headers = {
        "Content-Disposition": f'attachment; filename="{filename}"'
    }
    return StreamingResponse(
        out,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers
    )


@router.post("/{evaluation_id}/competencies/import")
async def import_competencies_xlsx(
    evaluation_id: UUID,
    file: UploadFile = File(...),
    current_user: CurrentUser = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> dict:
    _ensure_evaluation(db, current_user.company_id, evaluation_id)

    try:
        content = await file.read()
        wb = openpyxl.load_workbook(filename=BytesIO(content), data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="El archivo no es un Excel XLSX válido")

    ws = wb.active
    if not ws:
        raise HTTPException(status_code=400, detail="La hoja de cálculo está vacía")

    competencies_created = 0
    questions_created = 0
    questions_skipped = 0

    # Determine max position of current questions
    max_position = db.scalar(
        select(func.max(Question.position))
        .where(
            Question.company_id == current_user.company_id,
            Question.evaluation_id == evaluation_id,
        )
    ) or 0

    local_competency_cache = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue

        comp_name = str(row[0]).strip()
        if not comp_name:
            continue

        comp_desc = str(row[1]).strip() if row[1] is not None else None

        try:
            comp_weight = float(row[2]) if row[2] is not None else 1.0
        except (ValueError, TypeError):
            comp_weight = 1.0

        question_text = str(row[3]).strip() if row[3] is not None else None

        competency = local_competency_cache.get(comp_name)
        if not competency:
            competency = db.scalar(
                select(Competency).where(
                    Competency.company_id == current_user.company_id,
                    Competency.evaluation_id == evaluation_id,
                    Competency.name == comp_name,
                )
            )

        if not competency:
            competency = Competency(
                company_id=current_user.company_id,
                evaluation_id=evaluation_id,
                name=comp_name,
                description=comp_desc,
                weight=comp_weight,
            )
            db.add(competency)
            db.flush()
            competencies_created += 1
            local_competency_cache[comp_name] = competency

        if question_text:
            question_exists = db.scalar(
                select(Question).where(
                    Question.company_id == current_user.company_id,
                    Question.evaluation_id == evaluation_id,
                    Question.competency_id == competency.id,
                    Question.text == question_text,
                )
            )
            if question_exists:
                questions_skipped += 1
                continue

            # Parsing new columns
            question_type = "numeric_1_10"
            options = None
            is_evaluative = True

            if len(row) > 4 and row[4] is not None:
                question_type = str(row[4]).strip().lower()
            if len(row) > 5 and row[5] is not None:
                raw_opts = str(row[5]).strip()
                if raw_opts:
                    if question_type == "semantic_differential":
                        parts = [p.strip() for p in raw_opts.split("-") if p.strip()]
                        if len(parts) == 2:
                            options = {"left_label": parts[0], "right_label": parts[1], "steps": 7}
                        else:
                            parts = [p.strip() for p in raw_opts.split(",") if p.strip()]
                            if len(parts) == 2:
                                options = {"left_label": parts[0], "right_label": parts[1], "steps": 7}
                            else:
                                options = [p.strip() for p in raw_opts.split(",") if p.strip()]
                    else:
                        options = [p.strip() for p in raw_opts.split(",") if p.strip()]

            if len(row) > 6 and row[6] is not None:
                is_evaluative = str(row[6]).strip().upper() in ("SI", "YES", "TRUE", "1")

            max_position += 1
            new_question = Question(
                company_id=current_user.company_id,
                evaluation_id=evaluation_id,
                competency_id=competency.id,
                text=question_text,
                position=max_position,
                question_type=question_type,
                options=options,
                is_evaluative=is_evaluative,
            )
            db.add(new_question)
            questions_created += 1

    db.commit()

    return {
        "competencies_created": competencies_created,
        "questions_created": questions_created,
        "questions_skipped": questions_skipped,
    }
